import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

#inserts a formula in a cell
def insert_formula_cell(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path) #opens workbook
    sheet = workbook['População - Municipio'] #fill with desired sheet
    cell = sheet["A1"] #fill with desired cell
    cell.value = 'SUM(B1:B10)' #test formula
    workbook.save('your_excel_file_updated.xlsx') #saves workbook

#creates xlsx and fills it with values
def create_a_xlsx_and_edit(workbook_path):
    workbook = openpyxl.Workbook()

    data_sheet = workbook.create_sheet(title="data") #creates sheet called "data"
    data_sheet.title = "Database" #rename sheet
    resume_sheet = workbook.create_sheet(title="resume") #creates sheet called "resume"
    workbook.remove(workbook['Sheet']) #delete initial sheet
    #sellers
    data_sheet['A1'] = "sellers" #creates header
    data_sheet['A2'] = "Pedro" #seller1
    data_sheet['A3'] = "João" #seller2
    data_sheet['A4'] = "Luana" #seller3
    data_sheet['A5'] = "Mike" #seller4
    #amount sold
    data_sheet['B1'] = "sold" #creates header
    data_sheet['B2'] = 100 #sold1
    data_sheet['B3'] = 200 #sold2
    data_sheet['B4'] = 300 #sold3
    data_sheet['B5'] = 400 #sold4
    #sum values sold
    resume_sheet['A1'] = "Sum" #creates header
    resume_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center') #alignment adjusts
    resume_sheet['A1'].font = Font(bold=True) #font adjusts
    resume_sheet['A1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') #adjust color
    resume_sheet['A2'] = "=SUM(Database!B2:B5)" #sum values sold
    #save workbook
    workbook.save(workbook_path + ".xlsx") #save workbook

if __name__ == '__main__':
    workbook_path = input("whats the path to the workbook? ") #gets workbook_path
    #insert_formula_cell(workbook_path) #test function
    create_a_xlsx_and_edit(workbook_path) #test function
